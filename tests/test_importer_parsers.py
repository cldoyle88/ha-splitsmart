"""Parser tests — CSV against real fixtures, XLSX round-trip via openpyxl."""

from __future__ import annotations

import pathlib
from typing import Any

import openpyxl
import pytest

from custom_components.splitsmart.importer import csv_parser, xlsx_parser
from custom_components.splitsmart.importer.presets import (
    MONZO_MAPPING,
    REVOLUT_MAPPING,
    SPLITWISE_MAPPING,
    STARLING_MAPPING,
)

FIXTURES = pathlib.Path(__file__).parent / "fixtures" / "imports"


# --------------------------------------------------------- CSV inspect


@pytest.mark.asyncio
async def test_csv_inspect_detects_monzo() -> None:
    inspection = await csv_parser.inspect(FIXTURES / "monzo_classic.csv")
    assert inspection["preset"] == "Monzo"
    assert inspection["preset_confidence"] == "high"
    assert "Date" in inspection["headers"]
    assert len(inspection["sample_rows"]) == 10  # fixture has exactly 10 body rows
    assert inspection["file_origin_hash"].startswith("sha1:")


@pytest.mark.asyncio
async def test_csv_inspect_detects_starling() -> None:
    inspection = await csv_parser.inspect(FIXTURES / "starling_standard.csv")
    assert inspection["preset"] == "Starling"


@pytest.mark.asyncio
async def test_csv_inspect_detects_revolut() -> None:
    inspection = await csv_parser.inspect(FIXTURES / "revolut_account.csv")
    assert inspection["preset"] == "Revolut"


@pytest.mark.asyncio
async def test_csv_inspect_detects_splitwise() -> None:
    inspection = await csv_parser.inspect(FIXTURES / "splitwise_export.csv")
    assert inspection["preset"] == "Splitwise"


@pytest.mark.asyncio
async def test_csv_inspect_reports_no_preset_for_unknown_headers() -> None:
    inspection = await csv_parser.inspect(FIXTURES / "generic_no_preset.csv")
    assert inspection["preset"] is None
    assert inspection["preset_confidence"] is None


@pytest.mark.asyncio
async def test_csv_inspect_origin_hash_is_stable_across_calls() -> None:
    a = await csv_parser.inspect(FIXTURES / "monzo_classic.csv")
    b = await csv_parser.inspect(FIXTURES / "monzo_classic.csv")
    assert a["file_origin_hash"] == b["file_origin_hash"]


# --------------------------------------------------------- CSV parse


@pytest.mark.asyncio
async def test_csv_parse_monzo_happy_path() -> None:
    outcome = await csv_parser.parse(FIXTURES / "monzo_classic.csv", MONZO_MAPPING)
    assert outcome.row_count_raw == 10
    assert len(outcome.errors) == 0
    assert len(outcome.rows) == 10

    first = outcome.rows[0]
    assert first["date"] == "2026-04-15"
    assert first["description"] == "Waitrose"
    assert first["amount"] == 47.83  # Monzo exports negative; parser flips to positive
    assert first["currency"] == "GBP"
    assert first["category_hint"] == "Groceries"

    # The salary row is an income → stays negative after the sign flip.
    salary = next(r for r in outcome.rows if r["description"] == "Acme Ltd Salary")
    assert salary["amount"] == -2450.00


@pytest.mark.asyncio
async def test_csv_parse_starling_happy_path() -> None:
    outcome = await csv_parser.parse(FIXTURES / "starling_standard.csv", STARLING_MAPPING)
    assert outcome.row_count_raw == 10
    assert len(outcome.errors) == 0

    first = outcome.rows[0]
    assert first["date"] == "2026-04-15"
    assert first["description"] == "Waitrose"
    assert first["amount"] == 47.83
    assert first["currency"] == "GBP"
    assert first["category_hint"] == "Groceries"


@pytest.mark.asyncio
async def test_csv_parse_splitwise_happy_path() -> None:
    outcome = await csv_parser.parse(FIXTURES / "splitwise_export.csv", SPLITWISE_MAPPING)
    assert outcome.row_count_raw == 10
    assert len(outcome.errors) == 0

    first = outcome.rows[0]
    # Splitwise Cost is positive; expense_positive → amount stays positive.
    assert first["amount"] == 900.00
    assert first["category_hint"] == "Housing - Rent"


@pytest.mark.asyncio
async def test_csv_parse_revolut_captures_foreign_currencies() -> None:
    outcome = await csv_parser.parse(FIXTURES / "revolut_account.csv", REVOLUT_MAPPING)
    assert outcome.row_count_raw == 10
    assert len(outcome.errors) == 0

    currencies = {r["currency"] for r in outcome.rows}
    assert "EUR" in currencies
    assert "USD" in currencies
    assert "GBP" in currencies


@pytest.mark.asyncio
async def test_csv_parse_malformed_row_surfaces_as_error_not_abort() -> None:
    outcome = await csv_parser.parse(FIXTURES / "malformed.csv", MONZO_MAPPING)
    assert outcome.row_count_raw == 3
    assert len(outcome.rows) == 2  # two valid
    assert len(outcome.errors) == 1
    assert outcome.errors[0].row_index == 2  # second data row is malformed


@pytest.mark.asyncio
async def test_csv_parse_empty_file_returns_empty_outcome(tmp_path: pathlib.Path) -> None:
    empty = tmp_path / "empty.csv"
    empty.write_text("", encoding="utf-8")
    outcome = await csv_parser.parse(empty, MONZO_MAPPING)
    assert outcome.row_count_raw == 0
    assert outcome.rows == []


@pytest.mark.asyncio
async def test_csv_parser_handles_utf8_bom(tmp_path: pathlib.Path) -> None:
    path = tmp_path / "bom.csv"
    content = (
        "﻿Date,Name,Amount,Currency,Emoji,Notes and #tags,Description,Category\n"
        "15/04/2026,Café,-4.50,GBP,🥐,,Pastry,Eating out\n"
    )
    path.write_bytes(content.encode("utf-8"))
    outcome = await csv_parser.parse(path, MONZO_MAPPING)
    assert len(outcome.errors) == 0
    assert outcome.rows[0]["description"] == "Café"


@pytest.mark.asyncio
async def test_csv_parser_handles_cp1252_fallback(tmp_path: pathlib.Path) -> None:
    path = tmp_path / "cp1252.csv"
    content = (
        "Date,Name,Amount,Currency,Emoji,Notes and #tags,Description,Category\n"
        "15/04/2026,Caf\xe9,-4.50,GBP,,,Pastry,Eating out\n"
    )
    path.write_bytes(content.encode("cp1252"))
    outcome = await csv_parser.parse(path, MONZO_MAPPING)
    assert len(outcome.errors) == 0
    assert outcome.rows[0]["description"] == "Café"


# --------------------------------------------------------- XLSX


def _build_monzo_xlsx(path: pathlib.Path, rows: list[dict[str, Any]]) -> None:
    """Materialise an XLSX fixture at test time (no binary commit)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    wb.save(str(path))


@pytest.mark.asyncio
async def test_xlsx_inspect_and_parse_round_trip(tmp_path: pathlib.Path) -> None:
    path = tmp_path / "monzo.xlsx"
    _build_monzo_xlsx(
        path,
        [
            {
                "Date": "2026-04-15",
                "Name": "Waitrose",
                "Amount": -47.83,
                "Currency": "GBP",
                "Emoji": "🛒",
                "Notes and #tags": "",
                "Description": "Weekly shop",
                "Category": "Groceries",
            },
            {
                "Date": "2026-04-16",
                "Name": "TFL Travel",
                "Amount": -2.80,
                "Currency": "GBP",
                "Emoji": "🚇",
                "Notes and #tags": "",
                "Description": "Oyster",
                "Category": "Transport",
            },
        ],
    )
    inspection = await xlsx_parser.inspect(path)
    assert inspection["preset"] == "Monzo"

    outcome = await xlsx_parser.parse(path, MONZO_MAPPING)
    assert outcome.row_count_raw == 2
    assert len(outcome.errors) == 0
    assert outcome.rows[0]["description"] == "Waitrose"
    assert outcome.rows[0]["amount"] == 47.83


@pytest.mark.asyncio
async def test_xlsx_parser_handles_datetime_cells(tmp_path: pathlib.Path) -> None:
    """openpyxl returns dates as datetime objects; the parser must
    stringify them so apply_mapping's date parser picks them up."""
    import datetime as dt

    path = tmp_path / "dated.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "Date",
            "Name",
            "Amount",
            "Currency",
            "Emoji",
            "Notes and #tags",
            "Description",
            "Category",
        ]
    )
    ws.append(
        [dt.date(2026, 4, 15), "Waitrose", -47.83, "GBP", "🛒", "", "Weekly shop", "Groceries"]
    )
    wb.save(str(path))

    outcome = await xlsx_parser.parse(path, MONZO_MAPPING)
    assert outcome.rows[0]["date"] == "2026-04-15"


@pytest.mark.asyncio
async def test_xlsx_parser_tolerates_short_rows(tmp_path: pathlib.Path) -> None:
    """A row with fewer cells than headers (trailing empties elided) must
    not raise — empty cells fall through as empty strings."""
    path = tmp_path / "short.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "Date",
            "Name",
            "Amount",
            "Currency",
            "Emoji",
            "Notes and #tags",
            "Description",
            "Category",
        ]
    )
    ws.append(["2026-04-15", "Waitrose", -47.83, "GBP"])  # short: only 4 cells
    wb.save(str(path))

    outcome = await xlsx_parser.parse(path, MONZO_MAPPING)
    assert len(outcome.errors) == 0
    assert outcome.rows[0]["currency"] == "GBP"
